#python3 -m pip install openpyxl
from openpyxl import Workbook

# Luo uusi Excel-työkirja
wb = Workbook()

# Valitse aktiivinen taulukko
ws = wb.active
ws.title = "Esimerkki"

# Lisää otsikot
ws["A1"] = "Nimi"
ws["B1"] = "Ikä"

# Lisää dataa
ws.append(["Matti", 25])
ws.append(["Liisa", 30])
ws.append(["Pekka", 22])

# Tallenna tiedosto
wb.save("esimerkki.xlsx")

print("Excel-tiedosto luotu: esimerkki.xlsx")